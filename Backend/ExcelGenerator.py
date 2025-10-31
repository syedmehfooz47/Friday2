# -*- coding: utf-8 -*-
"""
Excel Generator - Creates professional Excel spreadsheets using Groq AI
Generates well-formatted, data-rich XLSX files
"""

import os
import time
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .logger import Logger
from .llm_handler import llm_handler

load_dotenv()


class ExcelGenerator:
    """Generate professional Excel spreadsheets"""
    
    def __init__(self):
        self.data_folder = Path(__file__).parent.parent / "Data" / "GeneratedDocuments"
        self.data_folder.mkdir(parents=True, exist_ok=True)
    
    def generate_excel(self, topic: str, rows: int = 20) -> tuple[str, str]:
        """
        Generate a professional Excel spreadsheet
        
        Args:
            topic: Topic/purpose for the spreadsheet
            rows: Approximate number of data rows to generate
            
        Returns:
            Tuple of (response message, file path)
        """
        username = os.getenv("Username", "Boss")
        
        if rows < 1:
            return f"Invalid row count, Boss. Please specify at least 1 row.", None
        
        Logger.log(f"Generating Excel spreadsheet with {rows} rows on topic: '{topic}'", "EXCEL")
        
        try:
            # Generate data structure using Groq
            content_prompt = f"""You are a professional data analyst. Create a well-structured Excel spreadsheet data about "{topic}".

Generate a table with:
- Appropriate column headers (5-8 columns)
- {rows} rows of realistic sample data
- Proper data types (text, numbers, dates where appropriate)

Format EXACTLY like this:

HEADERS:
Column1|Column2|Column3|Column4|Column5

DATA:
Value1|Value2|Value3|Value4|Value5
Value1|Value2|Value3|Value4|Value5

Make the data realistic and relevant to "{topic}"."""
            
            messages = [{"role": "user", "content": content_prompt}]
            content = llm_handler.get_response(messages, max_tokens=8000, temperature=0.7)
            
            if content.startswith("Error"):
                return f"Failed to generate content: {content}", None
            
            # Parse data
            sections = content.split('DATA:')
            
            if len(sections) < 2:
                return f"Failed to parse generated data, Boss.", None
            
            headers_section = sections[0].replace('HEADERS:', '').strip()
            data_section = sections[1].strip()
            
            headers = [h.strip() for h in headers_section.split('|')]
            
            data_rows = []
            for line in data_section.split('\n'):
                line = line.strip()
                if line and '|' in line:
                    row_data = [cell.strip() for cell in line.split('|')]
                    data_rows.append(row_data)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = topic[:31]  # Excel sheet name limit
            
            # Styling
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = 15
            
            # Add data
            for row_num, row_data in enumerate(data_rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    # Try to convert to number if possible
                    try:
                        cell.value = float(value)
                    except ValueError:
                        cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Save workbook
            safe_topic = "".join(c for c in topic if c.isalnum() or c in (' ', '_')).rstrip()[:50]
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            file_path = self.data_folder / f"{safe_topic.replace(' ', '_')}_{timestamp}.xlsx"
            
            wb.save(str(file_path))
            
            Logger.log(f"Excel spreadsheet generated successfully: {file_path}", "EXCEL")
            return f"I've generated an Excel spreadsheet with {len(data_rows)} rows on '{topic}' for you, Boss.", str(file_path)
        
        except Exception as e:
            Logger.log(f"Error generating Excel: {e}", "ERROR")
            return f"Failed to generate Excel: {e}", None


# Global instance
excel_generator = ExcelGenerator()